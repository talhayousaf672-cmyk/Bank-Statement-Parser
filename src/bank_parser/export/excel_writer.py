"""Excel export — turns a validated ParseResult into a styled XLSX workbook."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bank_parser.core.models import Language, ParseResult
from bank_parser.export.header_map import HEADER_MAP, HEADER_MAP_EN

# Languages that need right-to-left sheet direction
_RTL_LANGUAGES = {Language.ARABIC, Language.URDU}

# Header style constants
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_FLAG_FILL = PatternFill("solid", fgColor="C00000")
_FLAG_FONT = Font(bold=True, color="FFFFFF", size=11)


def write_excel(
    parse_result: ParseResult,
    output_path: str | Path,
    language: Language | None = None,
) -> Path:
    """Write a validated ParseResult to an XLSX workbook.

    Produces two sheets:
    - "Statement" — 16-field transaction rows with translated headers.
    - "Review Flags" — one row per review flag (statement-level and row-level).

    RTL sheet direction is enabled for Arabic and Urdu.
    """
    lang = language or parse_result.metadata.language
    headers = HEADER_MAP.get(lang, HEADER_MAP_EN)

    wb = openpyxl.Workbook()

    _write_statement_sheet(wb, parse_result, headers, lang)
    _write_review_flags_sheet(wb, parse_result)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def _write_statement_sheet(
    wb: openpyxl.Workbook,
    parse_result: ParseResult,
    headers: list[str],
    lang: Language,
) -> None:
    ws = wb.active
    ws.title = "Statement"

    if lang in _RTL_LANGUAGES:
        ws.sheet_view.rightToLeft = True

    # Write and style headers
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    meta = parse_result.metadata
    bank_label = _bank_label(meta.bank_id)
    for tx in parse_result.transactions:
        row = [
            bank_label,
            meta.account_number or "",
            meta.account_holder or "",
            str(meta.statement_period_start) if meta.statement_period_start else "",
            str(meta.statement_period_end) if meta.statement_period_end else "",
            meta.parser_version,
            meta.language.value,
            str(tx.transaction_date) if tx.transaction_date else "",
            str(tx.value_date) if tx.value_date else "",
            tx.description,
            tx.reference or "",
            float(tx.debit) if tx.debit is not None else "",
            float(tx.credit) if tx.credit is not None else "",
            float(tx.amount),
            float(tx.balance) if tx.balance is not None else "",
            tx.currency or meta.currency or "",
        ]
        ws.append(row)

    # Auto-size columns (cap at 50 chars)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


_BANK_ACRONYMS = {
    "hbl": "HBL",
    "ubl": "UBL",
    "mcb": "MCB",
    "mcb_bank": "MCB Bank",
}


def _bank_label(bank_id: str) -> str:
    if bank_id == "generic_english":
        return "AI Fallback"
    if bank_id.lower() in _BANK_ACRONYMS:
        return _BANK_ACRONYMS[bank_id.lower()]
    return bank_id.replace("_", " ").title()


def _write_review_flags_sheet(
    wb: openpyxl.Workbook,
    parse_result: ParseResult,
) -> None:
    ws = wb.create_sheet(title="Review Flags")

    flag_headers = ["Row", "Code", "Severity", "Message"]
    ws.append(flag_headers)
    for col_idx, _ in enumerate(flag_headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _FLAG_FONT
        cell.fill = _FLAG_FILL
        cell.alignment = Alignment(horizontal="center")

    # Statement-level flags (no row number)
    for flag in parse_result.review_flags:
        ws.append(["—", flag.code, flag.severity.value, flag.message])

    # Row-level flags
    for row_num, tx in enumerate(parse_result.transactions, start=1):
        for flag in tx.review_flags:
            ws.append([row_num, flag.code, flag.severity.value, flag.message])

    # Auto-size
    for col_idx in range(1, 5):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    if ws.max_row == 1:
        ws.append(["—", "—", "—", "No review flags — statement is clean."])
