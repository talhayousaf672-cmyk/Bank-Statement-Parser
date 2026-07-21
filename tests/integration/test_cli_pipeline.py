"""Integration tests — full end-to-end pipeline: fixture text → parse → validate → XLSX."""

from __future__ import annotations

from pathlib import Path
import json
from decimal import Decimal

import pytest
import openpyxl

from bank_parser.core.models import Language
from bank_parser.export.excel_writer import write_excel
from bank_parser.parsers.hbl import HBLParser
from bank_parser.parsers.ubl import UBLParser
from bank_parser.parsers.meezan_bank import MeezanBankParser
from bank_parser.parsers.wells_fargo import WellsFargoParser
from bank_parser.parsers.bank_of_america import BankOfAmericaParser
from bank_parser.parsers.chase_bank import ChaseBankParser
from bank_parser.validation.reconciliation import validate_parse_result
from bank_parser.validation.summary import summarize_validation

FIXTURE_BASE = Path("tests/fixtures/parsers")

_PARSER_MAP = {
    "hbl": (HBLParser, Language.ARABIC),
    "ubl": (UBLParser, Language.URDU),
    "meezan_bank": (MeezanBankParser, Language.ARABIC),
    "wells_fargo": (WellsFargoParser, Language.SPANISH),
    "bank_of_america": (BankOfAmericaParser, Language.RUSSIAN),
}


@pytest.mark.parametrize("bank_id,parser_lang", list(_PARSER_MAP.items()))
def test_full_pipeline_produces_xlsx(bank_id, parser_lang, tmp_path):
    """Full pipeline: fixture text → parse → validate → XLSX with correct structure."""
    ParserCls, lang = parser_lang
    text = (FIXTURE_BASE / bank_id / "en" / "statement_text.txt").read_text(encoding="utf-8")
    expected = json.loads((FIXTURE_BASE / bank_id / "en" / "expected_parse.json").read_text(encoding="utf-8"))

    # Parse + validate
    result = ParserCls().parse(text)
    result = validate_parse_result(result)

    # Summary: should be export ready (no errors in clean fixtures)
    summary = summarize_validation(result)
    assert summary.export_ready, (
        f"{bank_id}: expected export_ready but got {summary.export_readiness}. "
        f"Flags: {[f.code for f in result.review_flags]}"
    )

    # Export
    out_path = tmp_path / f"{bank_id}_output.xlsx"
    write_excel(result, out_path, language=lang)
    assert out_path.exists()

    # Verify XLSX structure
    wb = openpyxl.load_workbook(out_path)
    assert "Statement" in wb.sheetnames
    assert "Review Flags" in wb.sheetnames

    ws = wb["Statement"]
    # 1 header row + N transaction rows
    assert ws.max_row == 1 + len(expected["transactions"])
    # 16 columns
    assert ws.max_column == 16


@pytest.mark.parametrize("bank_id,parser_lang", list(_PARSER_MAP.items()))
def test_full_pipeline_transaction_count_matches_fixture(bank_id, parser_lang, tmp_path):
    ParserCls, _ = parser_lang
    text = (FIXTURE_BASE / bank_id / "en" / "statement_text.txt").read_text(encoding="utf-8")
    expected = json.loads((FIXTURE_BASE / bank_id / "en" / "expected_parse.json").read_text(encoding="utf-8"))
    result = ParserCls().parse(text)
    result = validate_parse_result(result)
    assert len(result.transactions) == len(expected["transactions"])


@pytest.mark.parametrize("bank_id,parser_lang", list(_PARSER_MAP.items()))
def test_full_pipeline_rtl_for_arabic_urdu(bank_id, parser_lang, tmp_path):
    ParserCls, lang = parser_lang
    if lang not in (Language.ARABIC, Language.URDU):
        pytest.skip(f"RTL test only for Arabic/Urdu, skipping {bank_id}")
    text = (FIXTURE_BASE / bank_id / "en" / "statement_text.txt").read_text(encoding="utf-8")
    result = validate_parse_result(ParserCls().parse(text))
    out_path = tmp_path / "rtl_test.xlsx"
    write_excel(result, out_path, language=lang)
    wb = openpyxl.load_workbook(out_path)
    assert wb["Statement"].sheet_view.rightToLeft is True


def test_ai_fallback_gate_integration():
    """Verify AI gate correctly classifies a clean parse result as ACCEPTED."""
    from bank_parser.ai.fallback_gate import validate_ai_fallback_result, AiFallbackStatus
    text = (FIXTURE_BASE / "hbl" / "en" / "statement_text.txt").read_text(encoding="utf-8")
    result = validate_parse_result(HBLParser().parse(text))
    gate = validate_ai_fallback_result(result)
    assert gate.accepted
    assert gate.status in (AiFallbackStatus.ACCEPTED, AiFallbackStatus.ACCEPTED_WITH_REVIEW)


def test_enrichment_raises_without_api_key():
    from bank_parser.ai.enrichment import enrich_descriptions, EnrichmentUnavailableError
    text = (FIXTURE_BASE / "hbl" / "en" / "statement_text.txt").read_text(encoding="utf-8")
    result = validate_parse_result(HBLParser().parse(text))
    with pytest.raises(EnrichmentUnavailableError):
        enrich_descriptions(result, api_key=None)


def test_onboarding_assist_raises_without_api_key():
    from bank_parser.ai.onboarding_assist import draft_parser_from_sample, OnboardingUnavailableError
    with pytest.raises(OnboardingUnavailableError):
        draft_parser_from_sample("sample text", "test_bank", api_key=None)
