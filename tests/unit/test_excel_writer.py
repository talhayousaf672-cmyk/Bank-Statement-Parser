"""Tests for Excel export: header translation, RTL, 16 fields, review flags sheet."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from bank_parser.core.models import (
    Language,
    ParseResult,
    ReviewFlag,
    ReviewSeverity,
    StatementMetadata,
    Transaction,
)
from bank_parser.export.excel_writer import write_excel
from bank_parser.export.header_map import HEADER_MAP


def _make_result(language: Language = Language.SPANISH, *, with_flags: bool = False) -> ParseResult:
    flags = (
        [ReviewFlag(code="balance_mismatch", message="Mismatch.", severity=ReviewSeverity.ERROR, row_number=1)]
        if with_flags
        else []
    )
    tx = Transaction(
        description="Salary payment",
        amount=Decimal("1500.00"),
        credit=Decimal("1500.00"),
        balance=Decimal("3000.00"),
        currency="USD",
        review_flags=flags,
    )
    return ParseResult(
        metadata=StatementMetadata(
            bank_id="test_bank",
            language=language,
            account_number="12345678",
            account_holder="Umer Kashif",
            currency="USD",
        ),
        transactions=[tx],
    )


def test_write_excel_creates_file(tmp_path: Path) -> None:
    result = _make_result()
    out = write_excel(result, tmp_path / "out.xlsx")
    assert out.exists()
    assert out.suffix == ".xlsx"


def test_write_excel_produces_16_header_columns(tmp_path: Path) -> None:
    result = _make_result(Language.SPANISH)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Statement"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 17)]
    assert headers == HEADER_MAP[Language.SPANISH]


def test_write_excel_writes_transaction_row(tmp_path: Path) -> None:
    result = _make_result(Language.SPANISH)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Statement"]
    assert ws.max_row == 2  # 1 header + 1 transaction
    assert ws.cell(row=2, column=10).value == "Salary payment"  # description column
    assert ws.cell(row=2, column=14).value == 1500.0  # amount column


def test_write_excel_all_five_languages_produce_16_headers(tmp_path: Path) -> None:
    for lang in Language:
        result = _make_result(lang)
        out = write_excel(result, tmp_path / f"out_{lang.value}.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb["Statement"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 17)]
        assert len(headers) == 16, f"Language {lang} produced {len(headers)} headers"
        assert headers == HEADER_MAP[lang]


def test_write_excel_arabic_sheet_is_rtl(tmp_path: Path) -> None:
    result = _make_result(Language.ARABIC)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Statement"]
    assert ws.sheet_view.rightToLeft is True


def test_write_excel_urdu_sheet_is_rtl(tmp_path: Path) -> None:
    result = _make_result(Language.URDU)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Statement"]
    assert ws.sheet_view.rightToLeft is True


def test_write_excel_non_rtl_language_is_ltr(tmp_path: Path) -> None:
    result = _make_result(Language.RUSSIAN)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Statement"]
    assert not ws.sheet_view.rightToLeft


def test_write_excel_produces_review_flags_sheet(tmp_path: Path) -> None:
    result = _make_result(with_flags=True)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    assert "Review Flags" in wb.sheetnames


def test_write_excel_review_flags_sheet_has_flag_row(tmp_path: Path) -> None:
    result = _make_result(with_flags=True)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Review Flags"]
    # Row 1 = header, row 2 = flag
    codes = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "balance_mismatch" in codes


def test_write_excel_clean_statement_flags_sheet_says_clean(tmp_path: Path) -> None:
    result = _make_result(with_flags=False)
    out = write_excel(result, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["Review Flags"]
    # Only header row + the "no flags" placeholder row
    assert ws.max_row == 2
